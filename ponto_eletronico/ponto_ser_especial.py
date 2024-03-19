import pandas as pd
from openpyxl import load_workbook
import datetime


def is_valid_time(val):
    return isinstance(val, datetime.time)

# Função para processar cada aba (funcionário) e extrair os dados de ponto
def process_sheet(df):
    # Filtra as linhas que possuem entrada e saída válidas
    valid_entries = df.dropna(subset=['Entrada', 'Saída'], how='all')
    # Lista para armazenar as linhas do CSV
    csv_rows = []
    for index, row in valid_entries.iterrows():
        matricula = int(row['Matricula'])
        data_ponto = row['Data'].strftime('%d/%m/%Y')
        entrada = row['Entrada'] if pd.notnull(row['Entrada']) else None
        saida = row['Saída'] if pd.notnull(row['Saída']) else None

        if is_valid_time(entrada):
            entrada = str(entrada)
            entrada = entrada[:5]
            entrada = entrada.replace(":", "")
            csv_rows.append(f"{matricula},{data_ponto},{entrada},001")
        if is_valid_time(saida):
            saida = str(saida)
            saida = saida[:5]
            saida = saida.replace(":", "")
            csv_rows.append(f"{matricula},{data_ponto},{saida},001")
            
    return csv_rows

# Caminho para o arquivo Excel
excel_file_path = 'C:\Automacoes\Projetos_Syngenta\ponto_eletronico\ponto_ser_especial.xlsx'

# Carrega o arquivo Excel para obter os nomes das abas
workbook = load_workbook(filename=excel_file_path, data_only=True)

# Lista para armazenar todas as linhas CSV de todas as abas
all_csv_rows = []

# Itera por cada aba na planilha
for sheet_name in workbook.sheetnames:
    # Usa pandas para ler os dados da aba atual
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name, skiprows=5)  # Ajuste o número de linhas a pular se necessário
    # Processa a aba e adiciona as linhas ao resultado final
    all_csv_rows.extend(process_sheet(df))

# Converte as linhas CSV em um DataFrame
df_csv = pd.DataFrame(all_csv_rows)

# Salva o DataFrame em um arquivo .csv
df_csv.to_csv('pontos.csv', index=False, header=False)
